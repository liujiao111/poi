from urllib.parse import quote
from urllib import request
import json
import os
import threading
import random
import math
import pandas as pd
from transCoordinateSystem import gcj02_to_wgs84, gcj02_to_bd09
from shp import trans_point_to_shp
import xlwt
from openpyxl import Workbook

'''
版本更新说明：
2019.10.05：
    1. 数据导出格式支持CSV格式以及XLS两种格式;
    2. 支持同时采集多个城市的POI数据;
    3. 支持同时采集多个POI分类数据

2019.10.10:
    1. 数据导出支持CSV以及XLS两种格式;
    2. CSV格式数据会生成.shp文件，可以直接在ARCGIS中使用


'''

#################################################需要修改###########################################################

# TODO 1.替换为从高德开放平台上申请申请的密钥
amap_key_list = ['高德密钥1', '高德密钥2']

# TODO 2.城市，多个用逗号隔开
# city_list = ['']
city_list = ['衡阳', '武汉']

# TODO 3.分类关键字,最好对照<<高德地图POI分类关键字以及编码.xlsx>>来填写对应编码，多个用逗号隔开
# type_list = ['']
type_list = ['190307']

# TODO 4.搜索关键词，例如“农家乐”
keyword = ''
# keyword = '农家乐'

# TODO 5.输出数据坐标系,1为高德GCJ20坐标系，2WGS84坐标系，3百度BD09坐标系
coord = 2

# TODO 6. 输出数据文件格式,1为默认xls格式，2为csv格式，3为xlsx格式
data_file_format = 3

############################################以下不需要动#######################################################################


poi_search_url = "http://restapi.amap.com/v3/place/text"
poi_boundary_url = "https://ditu.amap.com/detail/get/detail"

poi_xingzheng_distrinct_url = "https://restapi.amap.com/v3/config/district?subdistrict=1&key=4188efb67360681f89110ccdb11e563b"


# 根据城市名称和分类关键字获取poi数据
def getpois(cityname, type, keyword):
    amap_key = random.sample(amap_key_list, 1)[0]
    poilist = []

    data = ''
    req_url = '%s?key=%s&extensions=all&city=%s&types=%s&keywords=%s&citylimit=true&offset=25&page=%s&output=json' \
              % (poi_search_url, amap_key, quote(cityname), quote(type), quote(keyword), 1)
    with request.urlopen(req_url) as f:
        data = f.read()
        data = data.decode('utf-8')

    resp = json.loads(data)  # 将字符串转换为json
    print(resp)

    total = int(resp.get('count', 0))
    poilist.extend(resp.get('pois', []))

    if total > 20:
        data = ''
        page = math.ceil(total / 20)
        for i in range(1, page):
            req_url = '%s?key=%s&extensions=all&city=%s&types=%s&keywords=%s&citylimit=true&offset=25&page=%s&output=json' \
                      % (poi_search_url, amap_key, quote(cityname), quote(type), quote(keyword), i+1)
            with request.urlopen(req_url) as f:
                data = f.read()
                data = data.decode('utf-8')
            resp = json.loads(data)  # 将字符串转换为json
            print(resp)
            poilist.extend(resp.get('pois', []))

    return poilist


# 数据写入excel
def write_to_excel(poilist, cityname, classfield):
    # 一个Workbook对象，这就相当于创建了一个Excel文件
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet(classfield, cell_overwrite_ok=True)

    # 第一行(列标题)
    sheet.write(0, 0, 'lon')
    sheet.write(0, 1, 'lat')
    sheet.write(0, 2, 'name')
    sheet.write(0, 3, 'address')
    sheet.write(0, 4, 'pname')
    sheet.write(0, 5, 'cityname')
    sheet.write(0, 6, 'business_area')
    sheet.write(0, 7, 'type')

    for i in range(len(poilist)):
        location = poilist[i]['location']
        name = poilist[i]['name']
        address = poilist[i]['address']
        pname = poilist[i]['pname']
        cityname = poilist[i]['cityname']
        business_area = poilist[i]['business_area']
        type = poilist[i]['type']
        lng = str(location).split(",")[0]
        lat = str(location).split(",")[1]

        if coord == 2:
            result = gcj02_to_wgs84(float(lng), float(lat))
            lng = result[0]
            lat = result[1]
        elif coord == 3:
            result = gcj02_to_bd09(float(lng), float(lat))
            lng = result[0]
            lat = result[1]

        # 每一行写入
        sheet.write(i + 1, 0, lng)
        sheet.write(i + 1, 1, lat)
        sheet.write(i + 1, 2, name)
        sheet.write(i + 1, 3, address)
        sheet.write(i + 1, 4, pname)
        sheet.write(i + 1, 5, cityname)
        sheet.write(i + 1, 6, business_area)
        sheet.write(i + 1, 7, type)

    # 最后，将以上操作保存到指定的Excel文件中
    book.save(r'data' + os.sep + 'poi-' + cityname + "-" + classfield + ".xls")


# 数据写入excel
def write_to_xlsx(poilist, cityname, classfield):
    # 一个Workbook对象，这就相当于创建了一个Excel文件
    wb = Workbook()
    ws = wb.active

    # 第一行(列标题)
    ws['A1'] = 'lon'
    ws['B1'] = 'lat'
    ws['C1'] = 'name'
    ws['D1'] = 'address'
    ws['E1'] = 'pname'
    ws['F1'] = 'cityname'
    ws['G1'] = 'business_area'
    ws['H1'] = 'type'

    for i, poi in enumerate(poilist):
        row = i + 2
        name = poi['name']
        address = poi['address']
        pname = poi['pname']
        cityname = poi['cityname']
        business_area = json.dumps(poi['business_area'], ensure_ascii=False)
        type = poi['type']
        lng = poi['location'].split(",")[0]
        lat = poi['location'].split(",")[1]

        if coord == 2:
            result = gcj02_to_wgs84(float(lng), float(lat))
            lng = result[0]
            lat = result[1]
        elif coord == 3:
            result = gcj02_to_bd09(float(lng), float(lat))
            lng = result[0]
            lat = result[1]

        # 每一行写入
        ws['A%s' % row] = lng
        ws['B%s' % row] = lat
        ws['C%s' % row] = name
        ws['D%s' % row] = address
        ws['E%s' % row] = pname
        ws['F%s' % row] = cityname
        ws['G%s' % row] = business_area
        ws['H%s' % row] = type

    # 最后，将以上操作保存到指定的Excel文件中
    folder = os.path.join(os.getcwd(), 'data')
    if not os.path.exists(folder):
        os.mkdir(folder)

    file_path = os.path.join(folder, 'poi-%s-%s.xlsx' % (cityname, classfield))
    wb.save(file_path)


# 数据写入csv文件中
def write_to_csv(poilist, cityname, classfield):
    data_csv = {}
    lons, lats, names, addresss, pnames, citynames, business_areas, types = [], [], [], [], [], [], [], []

    for i in range(len(poilist)):
        location = poilist[i]['location']
        name = poilist[i]['name']
        address = poilist[i]['address']
        pname = poilist[i]['pname']
        cityname = poilist[i]['cityname']
        business_area = poilist[i]['business_area']
        type = poilist[i]['type']
        lng = str(location).split(",")[0]
        lat = str(location).split(",")[1]

        if (coord == 2):
            result = gcj02_to_wgs84(float(lng), float(lat))
            lng = result[0]
            lat = result[1]
        if (coord == 3):
            result = gcj02_to_bd09(float(lng), float(lat))
            lng = result[0]
            lat = result[1]
        lons.append(lng)
        lats.append(lat)
        names.append(name)
        addresss.append(address)
        pnames.append(pname)
        citynames.append(cityname)
        if business_area == []:
            business_area = ''
        business_areas.append(business_area)
        types.append(type)
    data_csv['lon'], data_csv['lat'], data_csv['name'], data_csv['address'], data_csv['pname'], \
    data_csv['cityname'], data_csv['business_area'], data_csv['type'] = \
        lons, lats, names, addresss, pnames, citynames, business_areas, types

    df = pd.DataFrame(data_csv)

    folder_name = 'poi-' + cityname + "-" + classfield
    folder_name_full = 'data' + os.sep + folder_name + os.sep
    if os.path.exists(folder_name_full) is False:
        os.makedirs(folder_name_full)

    file_name = 'poi-' + cityname + "-" + classfield + ".csv"
    file_path = folder_name_full + file_name

    df.to_csv(file_path, index=False, encoding='utf_8_sig')
    return folder_name_full, file_name


def get_areas(code):
    '''
    获取城市的所有区域
    :param code:
    :return:
    '''

    print('获取城市的所有区域：code: ' + str(code).strip())
    data = get_distrinctNoCache(code)

    print('get_distrinct result:' + data)

    data = json.loads(data)

    districts = data['districts'][0]['districts']
    # 判断是否是直辖市
    # 北京市、上海市、天津市、重庆市。
    if code.startswith('重庆') or code.startswith('上海') or code.startswith('北京') or code.startswith('天津'):
        districts = data['districts'][0]['districts'][0]['districts']

    area = ','.join(district['adcode'] for district in districts)

    print(area)
    return area


def get_data(city, type, keyword):
    '''
    根据城市名以及POI类型爬取数据
    :param city:
    :param type:
    :param keyword:
    :return:
    '''

    area = ''
    isNeedAreas = True
    if isNeedAreas:
        area = get_areas(city)

    all_pois = []
    if area != None and area != "":
        area_list = str(area).split(",")
        if area_list == 0:
            area_list = str(area).split("，")

        for area in area_list:
            pois_area = getpois(area, type, keyword)
            print('当前城区：' + str(area) + ', 分类：' + str(type) + ", 总的有" + str(len(pois_area)) + "条数据")
            all_pois.extend(pois_area)
    else:
        all_pois = getpois(area, keyword)

    print("所有城区的数据汇总，总数为：" + str(len(all_pois)))

    if data_file_format == 2:
        # 写入CSV
        file_folder, file_name = write_to_csv(all_pois, city, keyword)
        # # 写入SHP
        # trans_point_to_shp(file_folder, file_name, 0, 1)
    elif data_file_format == 3:
        write_to_xlsx(all_pois, city, type)
    else:
        write_to_excel(all_pois, city, type)



def get_distrinctNoCache(code):
    '''
    获取中国城市行政区划
    :return:
    '''

    url = "https://restapi.amap.com/v3/config/district?subdistrict=2&extensions=all&key=4188efb67360681f89110ccdb11e563b"

    req_url = url + "&keywords=" + quote(code)

    print(req_url)

    with request.urlopen(req_url) as f:
        data = f.read()
        data = data.decode('utf-8')
    print(code, data)
    return data


if __name__ == '__main__':
    thread_list = []
    for city in city_list:
        for type in type_list:
            get_data_thread = threading.Thread(target=get_data, args=(city, type, keyword))
            thread_list.append(get_data_thread)

    for t in thread_list:
        t.setDaemon(True)
        t.start()
    for t in thread_list:
        t.join()

    print('总的', len(city_list), '个城市, ', len(type_list), '个分类数据全部爬取完成!')
